
from __future__ import annotations

import io
import shutil
import subprocess
import sys
import tempfile
from pathlib import Path

import pandas as pd
import streamlit as st
from openpyxl import load_workbook

APP_DIR = Path(__file__).resolve().parent
ASSETS_DIR = APP_DIR / "assets"
MATCHER_NAME = "Vehicle_Compatibility_Matcher_Focused_Report_Final.py"
WORKBOOK_NAME = "Vehicle_Compatibility_Workbench_Focused_Report_Final.xlsx"
REPORT_NAME = "Final_Report_Focused_Final.xlsx"

st.set_page_config(page_title="Fleet CAN Compatibility", layout="wide")


def write_inputs_to_workbook(workbook_path: Path, vehicles: list[str]) -> None:
    wb = load_workbook(workbook_path)
    ws = wb["Vehicle_Input"]
    max_clear = max(ws.max_row, len(vehicles) + 10, 1000)
    for r in range(2, max_clear + 1):
        ws[f"A{r}"] = None
        ws[f"B{r}"] = None
        ws[f"C{r}"] = None
        ws[f"D{r}"] = None
        ws[f"E{r}"] = None
    for idx, vehicle in enumerate(vehicles, start=2):
        ws[f"A{idx}"] = idx - 1
        ws[f"B{idx}"] = vehicle
        ws[f"C{idx}"] = ""
    wb.save(workbook_path)


def parse_vehicle_lines(raw_text: str) -> list[str]:
    raw = (raw_text or "").strip()
    raw = raw.replace("\u2028", "\n").replace("\u2029", "\n").replace("\r", "\n")
    lines = [" ".join(line.split()) for line in raw.split("\n") if line.strip()]
    vehicles: list[str] = []
    buffer = ""
    for line in lines:
        if line.isdigit() and len(line) == 4:
            if buffer:
                vehicles.append(f"{buffer} {line}".strip())
                buffer = ""
            elif vehicles:
                vehicles[-1] = f"{vehicles[-1]} {line}".strip()
            else:
                vehicles.append(line)
        elif len(line) >= 4 and line[-4:].isdigit():
            if buffer:
                vehicles.append(f"{buffer} {line}".strip())
                buffer = ""
            else:
                vehicles.append(line)
        else:
            buffer = f"{buffer} {line}".strip() if buffer else line
    if buffer:
        vehicles.append(buffer)
    return vehicles


def run_matcher(vehicles: list[str]) -> tuple[Path, Path]:
    temp_dir = Path(tempfile.mkdtemp(prefix="fleet_can_"))
    matcher_path = temp_dir / MATCHER_NAME
    workbook_path = temp_dir / WORKBOOK_NAME
    shutil.copy2(ASSETS_DIR / MATCHER_NAME, matcher_path)
    shutil.copy2(ASSETS_DIR / WORKBOOK_NAME, workbook_path)
    write_inputs_to_workbook(workbook_path, vehicles)
    result = subprocess.run(
        [sys.executable, str(matcher_path)],
        cwd=str(temp_dir),
        capture_output=True,
        text=True,
    )
    if result.returncode != 0:
        raise RuntimeError((result.stderr or result.stdout or "Matcher failed").strip())
    report_path = temp_dir / REPORT_NAME
    if not report_path.exists():
        raise RuntimeError("The matcher finished, but the final report file was not created.")
    return workbook_path, report_path


def read_kpis(report_path: Path) -> pd.DataFrame:
    try:
        df = pd.read_excel(report_path, sheet_name="Summary_KPI", header=None)
    except Exception:
        return pd.DataFrame(columns=["Metric", "Value"])
    rows = []
    for i in range(len(df)):
        a = df.iloc[i, 0] if df.shape[1] > 0 else None
        b = df.iloc[i, 1] if df.shape[1] > 1 else None
        if pd.isna(a) or a in ["Metric", "Vehicle Compatibility Summary", "Number of vehicles under each brand", "Recommendations for review accuracy"]:
            continue
        if isinstance(a, str) and a.startswith("•"):
            continue
        if pd.notna(a) and pd.notna(b):
            rows.append({"Metric": str(a), "Value": b})
    return pd.DataFrame(rows)


def read_sheet(report_path: Path, sheet_name: str) -> pd.DataFrame:
    try:
        return pd.read_excel(report_path, sheet_name=sheet_name)
    except Exception:
        return pd.DataFrame()


st.title("Fleet CAN Compatibility Checker")
st.caption("Paste one vehicle per line in the format: Brand Model Year of Make")

sample = "Toyota Yaris 2020\nMercedes-Benz Actros 2040 2017\nAshok Leyland Dost 2021"
raw_text = st.text_area(
    "Vehicle list",
    value=sample,
    height=220,
    help="Enter one vehicle per line. Example: Toyota Yaris 2020",
)

run = st.button("Run Compatibility Check", type="primary", use_container_width=True)

if run:
    vehicles = parse_vehicle_lines(raw_text)
    if not vehicles:
        st.error("Enter at least one vehicle, one per line.")
    else:
        with st.spinner("Running compatibility analysis..."):
            try:
                workbook_path, report_path = run_matcher(vehicles)
                kpi_df = read_kpis(report_path)
                supported_df = read_sheet(report_path, "Supported_Vehicles")
                unsupported_df = read_sheet(report_path, "Unsupported_Vehicles")

                st.success("Analysis complete.")

                if not kpi_df.empty:
                    st.subheader("Summary")
                    cols = st.columns(min(4, max(1, len(kpi_df))))
                    for i, row in enumerate(kpi_df.itertuples(index=False)):
                        cols[i % len(cols)].metric(str(row.Metric), str(row.Value))

                st.subheader("Supported Vehicles")
                if supported_df.empty:
                    st.info("No supported vehicles found.")
                else:
                    st.dataframe(supported_df, use_container_width=True, hide_index=True)

                st.subheader("Unsupported / Review Vehicles")
                if unsupported_df.empty:
                    st.info("No unsupported vehicles found.")
                else:
                    st.dataframe(unsupported_df, use_container_width=True, hide_index=True)

                with open(report_path, "rb") as f:
                    st.download_button(
                        "Download Final Report (Excel)",
                        data=f.read(),
                        file_name="Final_Report_Focused_Final.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                    )

                with open(workbook_path, "rb") as f:
                    st.download_button(
                        "Download Workbook with Results",
                        data=f.read(),
                        file_name="Vehicle_Compatibility_Workbench_Focused_Report_Final.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                    )

            except Exception as e:
                st.error(str(e))
