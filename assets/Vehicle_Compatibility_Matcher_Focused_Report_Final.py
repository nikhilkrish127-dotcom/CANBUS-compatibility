
from pathlib import Path
import re
from collections import defaultdict

import pandas as pd
from rapidfuzz import fuzz
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment

BASE = Path(__file__).resolve().parent
WORKBOOK = BASE / "Vehicle_Compatibility_Workbench_Focused_Report_Final.xlsx"
FINAL_REPORT_XLSX = BASE / "Final_Report_Focused_Final.xlsx"
FINAL_REPORT_CSV = BASE / "Final_Report_Focused_Supported_Final.csv"

GEN_RE = re.compile(r"\b(mp[1-6]|mk[1-6]|j[123]\d{2}|w\d{3}|g\d{2,3}|xzu[a-z0-9]+|axzh\d+|axza\d+|gsz\d+|vs30|w447|tq|tga|tgs|tgx|gmt1xx|spa|tnga-f|clar|ld)\b", re.I)

STOP_TOKENS = {
    "type","facelift","electric","hybrid","diesel","petrol","phev","ev","ice",
    "wagon","sedan","truck","trucks","van","bus","buses","rigid","tractor","head",
    "4x2","4x4","6x4","6x2","7.5","ton","sport","tourer","signature","limited",
    "luxury","premium","similar","or","and","car","cars","coach","minibus",
    "pickup","pick","up","suv","hatchback","saloon","vehicle","asset","description",
    "passenger","staff","school","cargo"
}
CATEGORY_WORDS = {
    "bus","buses","coach","minibus","truck","trucks","tractor","rigid","tipper","hauler","lorry",
    "van","minivan","panel","car","cars","sedan","saloon","wagon","suv","hatchback","coupe","convertible","estate",
    "pickup","pick","up","cab","doublecab","double","crewcab","crew"
}
GENERIC_FAMILY_TERMS = {
    "", "bus", "truck", "van", "car", "pickup", "coach", "minibus", "staff bus", "school bus"
}

def norm_text(s):
    if pd.isna(s):
        return ""
    s = str(s).strip().lower()
    s = s.replace("&", " and ")
    s = re.sub(r"[\(\)\[\]\{\}/\\,._\-]+", " ", s)
    s = re.sub(r"[^a-z0-9+\s]", "", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s

def norm_make(s):
    s = norm_text(s)
    repl = {
        "mercedes benz":"mercedesbenz","mercedes":"mercedesbenz","benz":"mercedesbenz",
        "nissan ud":"ud","ud trucks":"ud","land rover":"landrover","great wall":"greatwall",
        "alfa romeo":"alfaromeo","rolls royce":"rollsroyce","vw":"volkswagen","chevy":"chevrolet",
        "ashok leyland":"ashokleyland","ashoka leyland":"ashokleyland",
        "bharat benz":"bharatbenz","force motors":"forcemotors","general motors":"gm",
    }
    for k, v in repl.items():
        s = s.replace(k, v)
    return s.replace(" ", "")

def extract_year(s):
    if pd.isna(s):
        return None
    m = re.search(r"(19|20)\d{2}", str(s))
    return int(m.group(0)) if m else None

def extract_generation(text):
    txt = norm_text(text)
    hits = GEN_RE.findall(txt)
    return ", ".join(sorted(set([h.lower() for h in hits]))) if hits else ""

def parse_year_ranges(year_text):
    if pd.isna(year_text):
        return []
    s = str(year_text).replace("–","-").replace("—","-").replace(" ","")
    s = s.replace(">=", "+").replace(">", "+")
    ranges = []
    for part in re.split(r"[;,/]", s):
        if not part:
            continue
        m = re.fullmatch(r"(\d{4})\+", part)
        if m:
            ranges.append((int(m.group(1)), None)); continue
        m = re.fullmatch(r"(\d{4})-(\d{4})", part)
        if m:
            ranges.append((int(m.group(1)), int(m.group(2)))); continue
        m = re.fullmatch(r"(\d{4})", part)
        if m:
            y = int(m.group(1)); ranges.append((y, y)); continue
        m = re.search(r"(\d{4})-(\d{4})", part)
        if m:
            ranges.append((int(m.group(1)), int(m.group(2)))); continue
        m = re.search(r"(\d{4})\+", part)
        if m:
            ranges.append((int(m.group(1)), None)); continue
    return ranges

def year_in_range(year, year_text):
    if year is None:
        return None
    ranges = parse_year_ranges(year_text)
    if not ranges:
        return None
    for start, end in ranges:
        if end is None and year >= start:
            return True
        if end is not None and start <= year <= end:
            return True
    return False

def year_distance(year, year_text):
    if year is None:
        return None
    ranges = parse_year_ranges(year_text)
    if not ranges:
        return None
    distances = []
    for start, end in ranges:
        if end is None:
            distances.append(0 if year >= start else start - year)
        else:
            if start <= year <= end:
                distances.append(0)
            elif year < start:
                distances.append(start - year)
            else:
                distances.append(year - end)
    return min(distances) if distances else None


def assess_year_match(year, year_text):
    """
    Returns a dict with:
      kind: Exact range | Narrow open range | Broad open range | Out of range | No list year | Not provided
      points: score contribution
      rank: tie-breaker rank (higher is better)
      distance: years away when out of range
    """
    if year is None:
        return {"kind": "Not provided", "points": 0, "rank": 0, "distance": None}
    ranges = parse_year_ranges(year_text)
    if not ranges:
        return {"kind": "No list year", "points": 0, "rank": 0, "distance": None}

    in_range_matches = []
    out_dists = []
    for start, end in ranges:
        if end is None:
            if year >= start:
                age = year - start
                if age <= 2:
                    kind = "Narrow open range"
                    points = 24
                    rank = 120 - age
                elif age <= 6:
                    kind = "Narrow open range"
                    points = 18
                    rank = 100 - age
                else:
                    kind = "Broad open range"
                    points = 4
                    rank = 40 - min(age, 30)
                in_range_matches.append((rank, {"kind": kind, "points": points, "rank": rank, "distance": 0}))
            else:
                out_dists.append(start - year)
        else:
            if start <= year <= end:
                width = max(1, end - start)
                rank = 140 - min(width, 30)
                points = 26 if width <= 4 else (22 if width <= 8 else 18)
                in_range_matches.append((rank, {"kind": "Exact range", "points": points, "rank": rank, "distance": 0}))
            elif year < start:
                out_dists.append(start - year)
            else:
                out_dists.append(year - end)

    if in_range_matches:
        return sorted(in_range_matches, key=lambda x: x[0], reverse=True)[0][1]
    distance = min(out_dists) if out_dists else None
    points = -12 if distance == 1 else (-18 if distance and distance <= 3 else -28)
    return {"kind": f"Out of range ({distance}y)" if distance is not None else "Out of range", "points": points, "rank": -distance if distance is not None else -99, "distance": distance}

def normalize_category(val):
    txt = norm_text(val)
    if not txt:
        return ""
    if any(t in txt for t in ["bus","buses","coach","minibus"]):
        return "bus"
    if any(t in txt for t in ["truck","trucks","tractor","rigid","tipper","hauler","lorry"]):
        return "truck"
    if any(t in txt for t in ["van","minivan","panel"]):
        return "van"
    if any(t in txt for t in ["pickup","double cab","doublecab","crew cab","crewcab"]):
        return "pickup"
    if any(t in txt for t in ["car","cars","sedan","saloon","wagon","suv","hatchback","coupe","convertible","estate"]):
        return "car"
    if any(t in txt for t in ["motorbike","motorcycle"]):
        return "motorbike"
    if any(t in txt for t in ["agricultural"]):
        return "agri"
    if any(t in txt for t in ["machinery","construction","utility","forest","heavy"]):
        return "machinery"
    if any(t in txt for t in ["motorboats","jet ski","snowmobile","marine","boat"]):
        return "marine"
    return ""

def infer_input_category(raw_desc):
    return normalize_category(raw_desc)

def infer_model_family(model_text):
    txt = norm_text(model_text)
    if not txt:
        return ""
    toks = [t for t in re.split(r"\s+", txt) if t]
    kept = []
    for i, t in enumerate(toks):
        if re.fullmatch(r"(19|20)\d{2}\+?", t):
            continue
        if t in STOP_TOKENS:
            continue
        if re.fullmatch(r"mp[1-6]|mk[1-6]|j[123]\d{2}|w\d{3}|g\d{2,3}|vs30|w447|tq|tga|tgs|tgx|gmt1xx|spa|clar|ld", t):
            continue
        if re.fullmatch(r"\d+", t):
            if i == 0:
                kept.append(t)
            continue
        kept.append(t)
        if len(kept) >= 3:
            break
    return " ".join(kept[:3]).strip()

def detect_fuel_support(param_text):
    if pd.isna(param_text):
        return "No"
    text = str(param_text).lower()
    fuel_keywords = [
        "fuel level","fuel level (%)","fuel level 1 (l)","fuel level 2 (l)","fuel level (l)",
        "fuel (l)","fuel volume","fuel consumption","instant fuel consumption",
        "average fuel consumption","fuel used","fuel rate","fuel flow","fuel cc",
    ]
    return "Yes" if any(k in text for k in fuel_keywords) else "No"

def reorder_parameters_for_display(param_text):
    if pd.isna(param_text):
        return ""
    parts = [p.strip() for p in str(param_text).split(";") if p.strip()]
    if not parts:
        return ""
    fuel_first, others = [], []
    for p in parts:
        (fuel_first if "fuel" in p.lower() else others).append(p)
    return "; ".join(fuel_first + others)

def autosize_sheet(ws):
    for col_cells in ws.columns:
        values = ["" if c.value is None else str(c.value) for c in col_cells]
        max_len = min(60, max((len(v) for v in values), default=0) + 2)
        ws.column_dimensions[col_cells[0].column_letter].width = max(10, max_len)

def prepare_report_sheet(wb, name, headers):
    if name in wb.sheetnames:
        ws = wb[name]
        wb.remove(ws)
    ws = wb.create_sheet(name)
    for idx, h in enumerate(headers, start=1):
        ws.cell(1, idx).value = h
    return ws

def update_table_ref(ws, table_name):
    from openpyxl.utils import get_column_letter
    if table_name in ws.tables:
        ws.tables[table_name].ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"

def color_results(ws, status_col_letter):
    fills = {
        "Strong Match": PatternFill("solid", fgColor="E2F0D9"),
        "Possible Match": PatternFill("solid", fgColor="FFF2CC"),
        "Review Needed": PatternFill("solid", fgColor="FCE4D6"),
        "No Reliable Match": PatternFill("solid", fgColor="F4CCCC"),
    }
    for r in range(2, ws.max_row + 1):
        val = ws[f"{status_col_letter}{r}"].value
        if val in fills:
            for c in ws[r]:
                c.fill = fills[val]

def normalize_master(master):
    master = master.copy()
    needed = ["Brand","Model","Model_Family","Model_Year_Text","Support_List_Type","Source_Sheet",
              "Brand_Norm","Model_Norm","Family_Norm","Generation_Hint","Supported_Parameter_Count",
              "Supported_Parameters_Sample","Vehicle_Category"]
    for c in needed:
        if c not in master.columns:
            master[c] = ""
    master["Brand_Norm"] = master["Brand_Norm"].fillna("").replace("nan","")
    master.loc[master["Brand_Norm"]=="","Brand_Norm"] = master["Brand"].map(norm_make)
    master["Model_Norm"] = master["Model_Norm"].fillna("").replace("nan","")
    master.loc[master["Model_Norm"]=="","Model_Norm"] = master["Model"].map(norm_text)
    master["Family_Norm"] = master["Family_Norm"].fillna("").replace("nan","")
    master.loc[master["Family_Norm"]=="","Family_Norm"] = master["Model_Family"].map(norm_text)
    master["Generation_Hint"] = master["Generation_Hint"].fillna("").replace("nan","")
    master.loc[master["Generation_Hint"]=="","Generation_Hint"] = master["Model"].map(extract_generation)
    master["Category_Bucket"] = master["Vehicle_Category"].apply(normalize_category)
    master["Full_Supported_Parameters"] = master["Supported_Parameters_Sample"].fillna("").astype(str).apply(reorder_parameters_for_display)
    master["Param_Preview"] = master["Full_Supported_Parameters"].fillna("").astype(str).apply(lambda x: "; ".join([p.strip() for p in x.split(";") if p.strip()][:5]))
    master["Fuel_Data_Available"] = master["Supported_Parameters_Sample"].apply(detect_fuel_support)
    return master

def load_sheet_df(sheet_name):
    return pd.read_excel(WORKBOOK, sheet_name=sheet_name)

def build_alias_maps(alias_df):
    make_aliases = []
    model_aliases = []
    category_aliases = []
    if alias_df is None or alias_df.empty:
        return make_aliases, model_aliases, category_aliases
    adf = alias_df.copy()
    adf["Rule_Type"] = adf["Rule_Type"].astype(str)
    if "Priority" in adf.columns:
        adf["Priority"] = pd.to_numeric(adf["Priority"], errors="coerce").fillna(0)
    else:
        adf["Priority"] = 0
    adf = adf.sort_values(["Priority"], ascending=False)
    for _, r in adf.iterrows():
        rec = {
            "src": norm_text(r.get("Input_Text","")),
            "dst": norm_text(r.get("Normalized_Output","")),
            "priority": r.get("Priority",0),
            "notes": str(r.get("Notes","") or "")
        }
        rt = str(r.get("Rule_Type","")).strip()
        if rt == "Make_Alias":
            make_aliases.append(rec)
        elif rt == "Model_Alias":
            model_aliases.append(rec)
        elif rt == "Category_Alias":
            category_aliases.append(rec)
    return make_aliases, model_aliases, category_aliases

def detect_make(raw_norm, master, make_aliases):
    joined = raw_norm.replace(" ", "")
    joined_make = norm_make(raw_norm)
    for r in make_aliases:
        src = r["src"]
        dst = r["dst"]
        if src and src in raw_norm:
            return norm_make(dst)
    brand_map = master[["Brand","Brand_Norm"]].dropna().drop_duplicates()
    candidates = []
    for _, br in brand_map.iterrows():
        brand_label = norm_text(br["Brand"])
        brand_norm = str(br["Brand_Norm"])
        if not brand_norm:
            continue
        if brand_label and re.search(rf"\b{re.escape(brand_label)}\b", raw_norm):
            return brand_norm
        if brand_norm in joined or brand_norm in joined_make:
            return brand_norm
        candidates.append((brand_label, brand_norm))
    # fuzzy brand rescue for typos like Ashoka Leyland
    prefix_tokens = []
    for tok in raw_norm.split():
        if re.fullmatch(r"(19|20)\d{2}", tok):
            continue
        if tok in CATEGORY_WORDS:
            continue
        prefix_tokens.append(tok)
        if len(prefix_tokens) >= 3:
            break
    prefix = " ".join(prefix_tokens)
    prefix_norm = norm_make(prefix)
    if not prefix:
        return ""
    scored = []
    for brand_label, brand_norm in candidates:
        score = max(fuzz.ratio(prefix, brand_label), fuzz.ratio(prefix_norm, brand_norm))
        if brand_label and prefix.startswith(brand_label):
            score += 8
        scored.append((score, brand_norm))
    scored.sort(reverse=True)
    if scored and scored[0][0] >= 90 and (len(scored) == 1 or scored[0][0] - scored[1][0] >= 4):
        return scored[0][1]
    return ""

def remove_detected_brand(raw_desc, parsed_make, master):
    txt = norm_text(raw_desc)
    if not parsed_make:
        return txt
    brand_variants = master.loc[master["Brand_Norm"] == parsed_make, "Brand"].dropna().astype(str).map(norm_text).unique().tolist()
    brand_variants = [b for b in brand_variants if b]
    cleaned = txt
    for bv in sorted(set(brand_variants), key=len, reverse=True):
        cleaned = re.sub(rf"\b{re.escape(bv)}\b", " ", cleaned)

    # fuzzy prefix trim for variants/typos like Ashoka Leyland vs Ashok Leyland
    raw_tokens = txt.split()
    best_n = 0
    best_score = 0
    for bv in brand_variants:
        bv_tokens = bv.split()
        low = max(1, len(bv_tokens) - 1)
        high = min(len(raw_tokens), len(bv_tokens) + 1)
        for n in range(low, high + 1):
            prefix = " ".join(raw_tokens[:n])
            score = max(fuzz.token_set_ratio(prefix, bv), fuzz.ratio(norm_make(prefix), parsed_make))
            if score > best_score:
                best_score = score
                best_n = n
    if best_score >= 88 and best_n > 0:
        cleaned = " ".join(raw_tokens[best_n:])

    cleaned = re.sub(r"\s+", " ", cleaned).strip()
    return cleaned

def build_search_context(raw_desc, parsed_make, input_category, category_aliases=None):
    txt = norm_text(raw_desc)
    txt = re.sub(r"(19|20)\d{2}", " ", txt)
    if parsed_make:
        txt = re.sub(re.escape(parsed_make), " ", txt)
    for cat in ["bus","buses","coach","minibus","truck","trucks","tractor","rigid","tipper","lorry",
                "van","minivan","panel","car","cars","sedan","saloon","wagon","suv","hatchback","pickup","cab"]:
        txt = re.sub(rf"\b{re.escape(cat)}\b", " ", txt)
    if category_aliases:
        for r in category_aliases:
            if r["src"] and r["src"] in txt:
                txt = txt.replace(r["src"], " ")
    txt = re.sub(r"\s+", " ", txt).strip()
    return txt

def special_family_rules(raw_desc, parsed_make):
    txt = norm_text(raw_desc)
    if parsed_make == "toyota":
        if "land cruiser" in txt: return "land cruiser"
        if "corolla cross" in txt: return "corolla cross"
        if "corolla" in txt: return "corolla"
        if "yaris" in txt: return "yaris"
        if "hilux" in txt: return "hilux"
        if "coaster" in txt: return "coaster"
    if parsed_make == "mercedesbenz":
        if "actros" in txt: return "actros"
        if "sprinter" in txt: return "sprinter"
        if "vito" in txt: return "vito"
        if "citaro" in txt: return "citaro"
    if parsed_make == "ashokleyland":
        if "dost" in txt: return "dost"
        if "partner" in txt: return "partner"
        if "ecomet" in txt: return "ecomet"
        if "boss" in txt: return "boss"
        if "viking" in txt: return "viking"
        if "lynx" in txt: return "lynx"
        if "falcon" in txt: return "falcon"
        if "cheetah" in txt: return "cheetah"
        if "oyster" in txt: return "oyster"
    if parsed_make == "volvo":
        if "xc90" in txt: return "xc90"
        if "s90" in txt: return "s90"
        if re.search(r"\bfl\b", txt): return "fl"
    if parsed_make == "man":
        if "tgs" in txt: return "tgs"
        if "tga" in txt: return "tga"
        if "tgx" in txt: return "tgx"
    if parsed_make == "hino" and re.search(r"\bxzu", txt): return "xzu"
    return ""

def extract_parsed_model_family(raw_desc, parsed_make, master, model_aliases):
    special = special_family_rules(raw_desc, parsed_make)
    if special:
        return special
    raw_norm = norm_text(raw_desc)
    for r in model_aliases:
        src = r["src"]; dst = r["dst"]
        notes = r["notes"].lower()
        brand_ok = (not parsed_make) or (not notes) or (parsed_make in norm_make(notes)) or (parsed_make in notes.replace(" ",""))
        if src and src in raw_norm and dst and brand_ok:
            return dst
    temp = remove_detected_brand(raw_desc, parsed_make, master)
    temp = re.sub(r"(19|20)\d{2}", " ", temp)
    temp = re.sub(r"\s+", " ", temp).strip()
    if not temp:
        return ""
    fams = master.loc[master["Brand_Norm"] == parsed_make, "Family_Norm"].dropna().astype(str).unique().tolist() if parsed_make else master["Family_Norm"].dropna().astype(str).unique().tolist()
    fams = [f for f in fams if f and f != "nan"]
    best, best_score = "", -1
    for fam in fams:
        score = fuzz.token_set_ratio(temp, fam)
        if fam and fam in temp:
            score += 20
        if score > best_score:
            best, best_score = fam, score
    if best_score >= 72:
        return best
    inferred = infer_model_family(temp)
    if inferred and norm_make(inferred) == parsed_make:
        return ""
    return inferred

def check_override(raw_desc, overrides):
    raw_norm = norm_text(raw_desc)
    if "Active" not in overrides.columns:
        return None
    active = overrides[overrides["Active"].astype(str).str.strip().str.lower() == "yes"].copy()
    for _, r in active.iterrows():
        needle = norm_text(r.get("Raw_Match_Text", ""))
        if needle and needle in raw_norm:
            return r.to_dict()
    return None

def informative_tokens(text):
    toks = []
    for t in norm_text(text).split():
        if t in STOP_TOKENS or t in CATEGORY_WORDS:
            continue
        if re.fullmatch(r"(19|20)\d{2}", t):
            continue
        if len(t) <= 1:
            continue
        toks.append(t)
    return toks

def compute_model_similarity(search_text, parsed_family, row):
    row_family = str(row["Family_Norm"])
    row_model = str(row["Model_Norm"])
    info_tokens = set(informative_tokens(search_text or parsed_family))
    cand_tokens = set(informative_tokens(f"{row_family} {row_model}"))
    overlap = len(info_tokens & cand_tokens)
    coverage = overlap / max(1, len(info_tokens)) if info_tokens else 0
    fam_ratio = fuzz.token_set_ratio(parsed_family, row_family) if parsed_family else 0
    model_ratio = fuzz.token_set_ratio(search_text, row_model) if search_text else 0
    family_to_model = fuzz.token_set_ratio(search_text, row_family) if search_text else 0
    best_ratio = max(fam_ratio, model_ratio, family_to_model)

    score = 0
    model_check = "Weak"
    if parsed_family and parsed_family == row_family:
        score += 34; model_check = "Exact"
    elif parsed_family and (parsed_family in row_family or row_family in parsed_family):
        score += 24; model_check = "Close"
    elif coverage >= 0.80 and best_ratio >= 80:
        score += 22; model_check = "Close"
    elif coverage >= 0.55 and best_ratio >= 72:
        score += 14; model_check = "Family"
    elif best_ratio >= 78:
        score += 12; model_check = "Family"
    elif best_ratio >= 66:
        score += 4; model_check = "Weak"
    else:
        score -= 10; model_check = "Weak"

    if row_family and row_family in (search_text or ""):
        score += 8
    elif row_model and row_model in (search_text or ""):
        score += 6
    return round(score, 2), model_check, round(best_ratio, 1), round(coverage, 2)

def generic_input_flag(parsed_family, search_text):
    pf = norm_text(parsed_family)
    st = norm_text(search_text)
    info = informative_tokens(st)
    if pf in GENERIC_FAMILY_TERMS:
        return True
    if not pf and len(info) <= 1:
        return True
    if pf and pf in {"bus","truck","van","car","pickup"}:
        return True
    return False

def score_candidate(search_text, parsed_make, parsed_family, parsed_year, parsed_gen, input_category, row):
    row_brand_norm = str(row["Brand_Norm"])
    if not parsed_make or parsed_make != row_brand_norm:
        return None

    brand_check = "Exact"
    brand_points = 45

    row_cat = str(row.get("Category_Bucket","") or "")
    if input_category:
        if row_cat == input_category:
            category_check = "Exact"
            category_points = 14
        elif row_cat == "":
            category_check = "Unknown"
            category_points = 2
        else:
            category_check = "Mismatch"
            category_points = -18
    else:
        category_check = "Not provided"
        category_points = 3 if row_cat else 0

    model_points, model_check, best_ratio, coverage = compute_model_similarity(search_text, parsed_family, row)

    gen_points = 0
    row_gen = norm_text(row.get("Generation_Hint",""))
    if parsed_gen:
        if row_gen and any(code.strip() in row_gen for code in parsed_gen.split(",")):
            gen_points = 15
        elif row_gen:
            gen_points = -20
    else:
        if row_gen:
            gen_points = -5

    year_eval = assess_year_match(parsed_year, row.get("Model_Year_Text", ""))
    year_check = year_eval["kind"]
    year_points = year_eval["points"]
    year_rank = year_eval["rank"]

    # If model is generic and year match is only broad open-ended, be conservative.
    broad_penalty = 0
    if year_check == "Broad open range" and model_check in {"Family", "Weak"}:
        broad_penalty = -12

    score = round(brand_points + category_points + model_points + gen_points + year_points + broad_penalty, 2)
    return {
        "score": score,
        "brand_check": brand_check,
        "model_check": model_check,
        "year_check": year_check,
        "category_check": category_check,
        "best_ratio": best_ratio,
        "coverage": coverage,
        "year_rank": year_rank,
        "year_distance": year_eval["distance"],
    }

def classify_candidate(cand, rows_for_brand, parsed_family, search_text):
    status = "Review Needed"
    confidence = "Low"
    if not cand:
        return "No Reliable Match", "Low"
    if cand["category_check"] == "Mismatch":
        return "No Reliable Match", "Low"

    generic = generic_input_flag(parsed_family, search_text)
    distinct_families = rows_for_brand["Family_Norm"].replace("", pd.NA).dropna().nunique() if isinstance(rows_for_brand, pd.DataFrame) and not rows_for_brand.empty else 0

    if cand["year_check"].startswith("Out of range"):
        return "No Reliable Match", "Low"
    if cand["year_check"] == "Broad open range" and not parsed_family:
        return "Review Needed", "Low"

    if cand["score"] >= 88 and cand["model_check"] in {"Exact","Close"} and cand["year_check"] in {"Exact range", "Narrow open range", "Not provided", "No list year"}:
        status, confidence = "Strong Match", "High"
    elif cand["score"] >= 72 and cand["model_check"] in {"Exact","Close","Family"} and cand["year_check"] in {"Exact range", "Narrow open range", "Broad open range", "Not provided", "No list year"}:
        status, confidence = "Possible Match", "Medium"
    elif cand["score"] >= 60 and cand["model_check"] != "Weak" and cand["category_check"] != "Mismatch":
        status, confidence = "Review Needed", "Medium"
    else:
        status, confidence = "No Reliable Match", "Low"

    if generic and distinct_families > 1 and status in {"Possible Match","Review Needed"} and cand["score"] < 84:
        status, confidence = "Review Needed", "Low"
    if cand["year_check"] == "Broad open range" and status == "Strong Match":
        status, confidence = "Possible Match", "Medium"
    return status, confidence

def overall_reason(brand_check, model_check, year_check, category_check):
    return f"Brand: {brand_check}; Category: {category_check}; Model: {model_check}; Year: {year_check}"

def action_hint(status, parsed_year):
    if status == "Strong Match":
        return "Auto-accept candidate"
    if status == "Possible Match":
        return "Review top shortlist before approving"
    if status == "Review Needed" and parsed_year is None:
        return "Ask for exact model or year"
    if status == "Review Needed":
        return "Manual review recommended"
    return "No reliable match - collect more specific model details"

def prepare_candidate_pool(rows, input_category, search_text):
    if rows.empty:
        return rows
    # quick prefilter for performance and precision
    if input_category:
        same_cat = rows[rows["Category_Bucket"] == input_category].copy()
        if not same_cat.empty:
            rows = same_cat
    rows = rows.copy()
    rows["_prefilter_ratio"] = rows["Family_Norm"].fillna("").astype(str).apply(lambda x: fuzz.token_set_ratio(search_text, x) if search_text else 0)
    rows["_prefilter_ratio2"] = rows["Model_Norm"].fillna("").astype(str).apply(lambda x: fuzz.token_set_ratio(search_text, x) if search_text else 0)
    rows["_prefilter"] = rows[["_prefilter_ratio","_prefilter_ratio2"]].max(axis=1)
    if len(rows) > 250:
        rows = rows.sort_values(["_prefilter"], ascending=False).head(250).copy()
    return rows

def build_focused_report(records, input_rows):
    def normalize_brand(val):
        val = str(val or "").strip()
        return val if val else "Unrecognized"

    def review_note(status, confidence, fuel_support, parsed_model, reason):
        notes = []
        if status in ["Review Needed", "No Reliable Match"]:
            notes.append("Review model wording against source support lists before relying on the result.")
        if confidence == "Low":
            notes.append("Low-confidence match; manual validation is recommended.")
        if not parsed_model or str(parsed_model).strip() in GENERIC_FAMILY_TERMS:
            notes.append("Input model wording is generic; exact supported model may need confirmation.")
        if fuel_support == "No":
            notes.append("Compatibility may exist, but fuel level is not confirmed readable from matched devices.")
        if "Out of range" in str(reason):
            notes.append("Check whether the vehicle year is outside supported ranges.")
        return " | ".join(dict.fromkeys(notes)) if notes else "No immediate review needed."

    summary_brand_df = pd.DataFrame(columns=["Brand", "Vehicle_Count"])
    supported_df = pd.DataFrame(columns=[
        "Brand", "Vehicle_Model", "Year_of_Make", "Compatible_CAN_Devices",
        "Fuel_Level_Reading", "Tested_Readable_Parameters", "Matched_Supported_Model(s)",
        "Review_Recommendation"
    ])
    unsupported_df = pd.DataFrame(columns=[
        "Brand", "Vehicle_Model", "Year_of_Make", "Why_Considered_Unsupported",
        "Review_Recommendation"
    ])

    if not records:
        metrics = {
            "Total number of vehicles": 0,
            "Total number of brands": 0,
            "Total number of unique models": 0,
            "Vehicles compatible with at least one CAN device": 0,
            "Compatible vehicles without fuel reading": 0,
            "Vehicles not or suspected to not be supported by any CAN device": 0,
            "Support coverage %": 0.0,
        }
        return metrics, summary_brand_df, supported_df, unsupported_df, ["No vehicle rows were available for analysis."]

    df = pd.DataFrame(records).copy()
    counts = pd.Series(input_rows).value_counts()

    grouped = []
    for vehicle, g in df.groupby("Raw_Vehicle_Description", sort=True):
        input_count = int(counts.get(vehicle, len(g)))
        g = g.sort_values(["Match_Score"], ascending=False).copy()
        supported = g[g["Status"].isin(["Strong Match", "Possible Match"])].copy()
        best_any = g.iloc[0]

        parsed_brand = normalize_brand(best_any.get("Parsed_Brand", ""))
        parsed_model = str(best_any.get("Parsed_Model_Family", "") or "").strip()
        parsed_year = best_any.get("Parsed_Year", "")
        if pd.isna(parsed_year):
            parsed_year = ""

        if not supported.empty:
            device_names = sorted({str(x).strip() for x in supported["Source"].tolist() if str(x).strip()})
            fuel_yes = supported["Fuel_Data_Available"].astype(str).str.upper().eq("YES")
            fuel_support = "Yes" if fuel_yes.any() else "No"
            matched_models = sorted({f'{str(r.get("Matched_Brand","")).strip()} {str(r.get("Matched_Model","")).strip()}'.strip()
                                     for _, r in supported.iterrows() if str(r.get("Matched_Model","")).strip()})
            device_param_parts = []
            for device, dg in supported.groupby("Source", sort=True):
                params = []
                for p in dg["Full_Supported_Parameters"].fillna("").astype(str).tolist():
                    params.extend([x.strip() for x in p.split(";") if x.strip()])
                params = list(dict.fromkeys(params))
                if not params:
                    preview_parts = []
                    for p in dg["Parameter_Preview"].fillna("").astype(str).tolist():
                        preview_parts.extend([x.strip() for x in p.split(";") if x.strip()])
                    params = list(dict.fromkeys(preview_parts))
                if params:
                    device_param_parts.append(f"{device}: " + "; ".join(params[:20]))
                else:
                    device_param_parts.append(f"{device}: No tested readable parameters listed")
            grouped.append({
                "Vehicle_Input": vehicle,
                "Input_Count": input_count,
                "Brand": parsed_brand,
                "Parsed_Model_Family": parsed_model,
                "Year_of_Make": parsed_year,
                "Supported": "Yes",
                "Best_Status": best_any.get("Status",""),
                "Confidence": best_any.get("Confidence",""),
                "Fuel_Level_Reading": fuel_support,
                "Compatible_CAN_Devices": ", ".join(device_names),
                "Tested_Readable_Parameters": " | ".join(device_param_parts),
                "Matched_Supported_Model(s)": " | ".join(matched_models[:10]),
                "Reason": str(best_any.get("Reason","")),
                "Review_Recommendation": review_note(best_any.get("Status",""), best_any.get("Confidence",""), fuel_support, parsed_model, str(best_any.get("Reason","")))
            })
        else:
            grouped.append({
                "Vehicle_Input": vehicle,
                "Input_Count": input_count,
                "Brand": parsed_brand,
                "Parsed_Model_Family": parsed_model,
                "Year_of_Make": parsed_year,
                "Supported": "No",
                "Best_Status": best_any.get("Status",""),
                "Confidence": best_any.get("Confidence",""),
                "Fuel_Level_Reading": "No",
                "Compatible_CAN_Devices": "",
                "Tested_Readable_Parameters": "",
                "Matched_Supported_Model(s)": f'{str(best_any.get("Matched_Brand","")).strip()} {str(best_any.get("Matched_Model","")).strip()}'.strip(),
                "Reason": str(best_any.get("Reason","")),
                "Review_Recommendation": review_note(best_any.get("Status",""), best_any.get("Confidence",""), "No", parsed_model, str(best_any.get("Reason","")))
            })

    final_df = pd.DataFrame(grouped)
    total_rows = int(final_df["Input_Count"].sum()) if not final_df.empty else 0
    brand_counts = final_df.groupby("Brand", dropna=False)["Input_Count"].sum().reset_index().rename(columns={"Input_Count":"Vehicle_Count"}).sort_values(["Vehicle_Count", "Brand"], ascending=[False, True])
    compatible_rows = final_df[final_df["Supported"] == "Yes"].copy()
    unsupported_rows = final_df[final_df["Supported"] == "No"].copy()

    compatible_vehicle_count = int(compatible_rows["Input_Count"].sum()) if not compatible_rows.empty else 0
    no_fuel_vehicle_count = int(compatible_rows.loc[compatible_rows["Fuel_Level_Reading"] == "No", "Input_Count"].sum()) if not compatible_rows.empty else 0
    unsupported_vehicle_count = int(unsupported_rows["Input_Count"].sum()) if not unsupported_rows.empty else 0

    metrics = {
        "Total number of vehicles": total_rows,
        "Total number of brands": int(brand_counts.loc[brand_counts["Brand"] != "Unrecognized", "Brand"].nunique()),
        "Total number of unique models": int(final_df["Vehicle_Input"].nunique()) if not final_df.empty else 0,
        "Vehicles compatible with at least one CAN device": compatible_vehicle_count,
        "Compatible vehicles without fuel reading": no_fuel_vehicle_count,
        "Vehicles not or suspected to not be supported by any CAN device": unsupported_vehicle_count,
        "Support coverage %": round((compatible_vehicle_count / total_rows) * 100, 1) if total_rows else 0.0,
    }

    supported_df = compatible_rows[[
        "Brand", "Vehicle_Input", "Year_of_Make", "Compatible_CAN_Devices",
        "Fuel_Level_Reading", "Tested_Readable_Parameters", "Matched_Supported_Model(s)",
        "Review_Recommendation"
    ]].copy().rename(columns={"Vehicle_Input":"Vehicle_Model"}).sort_values(["Brand", "Vehicle_Model", "Year_of_Make"]).reset_index(drop=True)

    unsupported_df = unsupported_rows[[
        "Brand", "Vehicle_Input", "Year_of_Make", "Reason", "Review_Recommendation"
    ]].copy().rename(columns={
        "Vehicle_Input": "Vehicle_Model",
        "Reason": "Why_Considered_Unsupported"
    }).sort_values(["Brand", "Vehicle_Model", "Year_of_Make"]).reset_index(drop=True)

    recommendations = []
    if unsupported_vehicle_count:
        recommendations.append(f"{unsupported_vehicle_count} fleet vehicles are not supported or are suspected to be unsupported; prioritize manual review of the Unsupported_Vehicles sheet.")
    if no_fuel_vehicle_count:
        recommendations.append(f"{no_fuel_vehicle_count} compatible fleet vehicles do not show fuel reading support; confirm whether fuel is required before device selection.")
    low_conf = int(final_df[(final_df["Best_Status"].isin(["Review Needed", "No Reliable Match"])) | (final_df["Confidence"] == "Low")]["Input_Count"].sum())
    if low_conf:
        recommendations.append(f"{low_conf} fleet vehicles carry review-needed or low-confidence outcomes; validate raw naming against source support lists for accuracy.")
    unrec = int(brand_counts.loc[brand_counts["Brand"] == "Unrecognized", "Vehicle_Count"].sum()) if not brand_counts.empty else 0
    if unrec:
        recommendations.append(f"{unrec} fleet vehicles have unrecognized brands; standardize brand names in the input list to improve matching accuracy.")
    if not recommendations:
        recommendations.append("No major review flags were found in the current list.")

    return metrics, brand_counts, supported_df, unsupported_df, recommendations


def write_focused_report_workbook(metrics, brand_counts, supported_df, unsupported_df, recommendations):
    from openpyxl import Workbook
    wb_rep = Workbook()
    ws_sum = wb_rep.active
    ws_sum.title = "Summary_KPI"

    header_fill = PatternFill("solid", fgColor="1F4E78")
    sub_fill = PatternFill("solid", fgColor="D9EAF7")
    white_font = Font(color="FFFFFF", bold=True)
    bold = Font(bold=True)

    ws_sum["A1"] = "Vehicle Compatibility Summary"
    ws_sum["A1"].font = Font(bold=True, size=14)
    ws_sum["A3"] = "Metric"
    ws_sum["B3"] = "Value"
    for c in ("A3", "B3"):
        ws_sum[c].fill = header_fill
        ws_sum[c].font = white_font

    row = 4
    for k, v in metrics.items():
        ws_sum.cell(row, 1).value = k
        ws_sum.cell(row, 2).value = v
        row += 1

    row += 2
    ws_sum.cell(row, 1).value = "Number of vehicles under each brand"
    ws_sum.cell(row, 1).fill = sub_fill
    ws_sum.cell(row, 1).font = bold
    row += 1
    ws_sum.cell(row, 1).value = "Brand"
    ws_sum.cell(row, 2).value = "Vehicle_Count"
    for c in ("A" + str(row), "B" + str(row)):
        ws_sum[c].fill = header_fill
        ws_sum[c].font = white_font
    row += 1
    if brand_counts.empty:
        ws_sum.cell(row, 1).value = "None"
        row += 1
    else:
        for rec in brand_counts.itertuples(index=False):
            ws_sum.cell(row, 1).value = rec.Brand
            ws_sum.cell(row, 2).value = rec.Vehicle_Count
            row += 1

    row += 2
    ws_sum.cell(row, 1).value = "Recommendations for review accuracy"
    ws_sum.cell(row, 1).fill = sub_fill
    ws_sum.cell(row, 1).font = bold
    row += 1
    for note in recommendations:
        ws_sum.cell(row, 1).value = "• " + note
        row += 1

    def add_df_sheet(name, df):
        ws = wb_rep.create_sheet(name)
        for idx, h in enumerate(df.columns, start=1):
            ws.cell(1, idx).value = h
            ws.cell(1, idx).fill = header_fill
            ws.cell(1, idx).font = white_font
        for r_idx, rec in enumerate(df.itertuples(index=False), start=2):
            for c_idx, v in enumerate(rec, start=1):
                ws.cell(r_idx, c_idx).value = v
        ws.freeze_panes = "A2"
        for row_cells in ws.iter_rows():
            for c in row_cells:
                c.alignment = Alignment(vertical="top", wrap_text=True)
        autosize_sheet(ws)

    add_df_sheet("Supported_Vehicles", supported_df)
    add_df_sheet("Unsupported_Vehicles", unsupported_df)

    for row_cells in ws_sum.iter_rows():
        for c in row_cells:
            c.alignment = Alignment(vertical="top", wrap_text=True)
    autosize_sheet(ws_sum)
    wb_rep.save(FINAL_REPORT_XLSX)


def write_focused_report_sheets_to_workbook(wb, metrics, brand_counts, supported_df, unsupported_df, recommendations):
    header_fill = PatternFill("solid", fgColor="1F4E78")
    sub_fill = PatternFill("solid", fgColor="D9EAF7")
    white_font = Font(color="FFFFFF", bold=True)
    bold = Font(bold=True)

    for name in ["Report_Summary", "Adaptor_Detail", "Adaptor_Summary", "Brand_Summary",
                 "Summary_KPI", "Final_Report", "No_Fuel_Vehicles", "Brand_Gaps", "Needs_Review",
                 "Supported_Vehicles", "Unsupported_Vehicles"]:
        if name in wb.sheetnames:
            del wb[name]

    ws_sum = wb.create_sheet("Summary_KPI")
    ws_sup = wb.create_sheet("Supported_Vehicles")
    ws_unsup = wb.create_sheet("Unsupported_Vehicles")

    ws_sum["A1"] = "Vehicle Compatibility Summary"
    ws_sum["A1"].font = Font(bold=True, size=14)
    ws_sum["A3"] = "Metric"
    ws_sum["B3"] = "Value"
    for c in ("A3", "B3"):
        ws_sum[c].fill = header_fill
        ws_sum[c].font = white_font

    row = 4
    for k, v in metrics.items():
        ws_sum.cell(row, 1).value = k
        ws_sum.cell(row, 2).value = v
        row += 1

    row += 2
    ws_sum.cell(row, 1).value = "Number of vehicles under each brand"
    ws_sum.cell(row, 1).fill = sub_fill
    ws_sum.cell(row, 1).font = bold
    row += 1
    ws_sum.cell(row, 1).value = "Brand"
    ws_sum.cell(row, 2).value = "Vehicle_Count"
    for c in ("A" + str(row), "B" + str(row)):
        ws_sum[c].fill = header_fill
        ws_sum[c].font = white_font
    row += 1
    if brand_counts.empty:
        ws_sum.cell(row, 1).value = "None"
        row += 1
    else:
        for rec in brand_counts.itertuples(index=False):
            ws_sum.cell(row, 1).value = rec.Brand
            ws_sum.cell(row, 2).value = rec.Vehicle_Count
            row += 1

    row += 2
    ws_sum.cell(row, 1).value = "Recommendations for review accuracy"
    ws_sum.cell(row, 1).fill = sub_fill
    ws_sum.cell(row, 1).font = bold
    row += 1
    for note in recommendations:
        ws_sum.cell(row, 1).value = "• " + note
        row += 1

    for ws, df in [(ws_sup, supported_df), (ws_unsup, unsupported_df)]:
        for idx, h in enumerate(df.columns, start=1):
            ws.cell(1, idx).value = h
            ws.cell(1, idx).fill = header_fill
            ws.cell(1, idx).font = white_font
        for r_idx, rec in enumerate(df.itertuples(index=False), start=2):
            for c_idx, v in enumerate(rec, start=1):
                ws.cell(r_idx, c_idx).value = v
        ws.freeze_panes = "A2"

    for ws in [ws_sum, ws_sup, ws_unsup]:
        for row_cells in ws.iter_rows():
            for c in row_cells:
                c.alignment = Alignment(vertical="top", wrap_text=True)
        autosize_sheet(ws)


def main():
    wb = load_workbook(WORKBOOK)
    ws_in = wb["Vehicle_Input"]
    ws_match = wb["Match_Results"]
    ws_source = wb["Per_Source_Results"]
    ws_top3 = wb["Top3_Candidates"]

    master = normalize_master(load_sheet_df("Supported_Master"))
    alias_df = load_sheet_df("Aliases_Rules")
    overrides = load_sheet_df("Manual_Overrides")
    try:
        gen_rules = load_sheet_df("Generation_Rules")
    except Exception:
        gen_rules = pd.DataFrame()
    try:
        chassis_rules = load_sheet_df("Chassis_Rules")
    except Exception:
        chassis_rules = pd.DataFrame()

    make_aliases, model_aliases, category_aliases = build_alias_maps(alias_df)

    for ws in [ws_match, ws_source, ws_top3]:
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for c in row:
                c.value = None

    match_row_ptr = 2
    source_row_ptr = 2
    top_row_ptr = 2
    all_records = []
    input_rows = []

    brand_groups = {brand: g.copy() for brand, g in master.groupby("Brand_Norm", dropna=False)}

    for r in range(2, ws_in.max_row + 1):
        input_id = ws_in[f"A{r}"].value
        raw_desc = ws_in[f"B{r}"].value
        if raw_desc is None or str(raw_desc).strip() == "":
            continue

        raw_desc = str(raw_desc).strip()
        input_rows.append(raw_desc)

        override = check_override(raw_desc, overrides)
        parsed_year = extract_year(raw_desc)
        parsed_gen = extract_generation(raw_desc)
        parsed_make = detect_make(norm_text(raw_desc), master, make_aliases)
        input_category = infer_input_category(raw_desc)
        parsed_family = extract_parsed_model_family(raw_desc, parsed_make, master, model_aliases)
        search_text = build_search_context(raw_desc, parsed_make, input_category, category_aliases)
        if not search_text:
            search_text = parsed_family

        if override:
            rows = master[
                (master["Support_List_Type"].astype(str) == str(override.get("Override_Source",""))) &
                (master["Brand"].astype(str) == str(override.get("Override_Brand",""))) &
                (master["Model"].astype(str) == str(override.get("Override_Model","")))
            ].copy()
        else:
            rows = brand_groups.get(parsed_make, master.iloc[0:0].copy()) if parsed_make else master.iloc[0:0].copy()
            rows = prepare_candidate_pool(rows, input_category, search_text)

        scored = []
        for _, row in rows.iterrows():
            score_info = score_candidate(search_text, parsed_make, parsed_family, parsed_year, parsed_gen, input_category, row)
            if not score_info:
                continue
            status, conf = classify_candidate(score_info, rows, parsed_family, search_text)
            reason_detail = (
                f"Brand locked to '{parsed_make or ''}'. Input category='{input_category or ''}'. "
                f"Search text='{search_text or ''}'. Parsed family='{parsed_family or ''}'. "
                f"Compared against brand='{row.get('Brand','')}', model='{row.get('Model','')}', "
                f"category='{row.get('Vehicle_Category','')}', year range='{row.get('Model_Year_Text','')}'. "
                f"Checks -> brand={score_info['brand_check']}, category={score_info['category_check']}, "
                f"model={score_info['model_check']}, year={score_info['year_check']}; "
                f"text_ratio={score_info['best_ratio']}, token_coverage={score_info['coverage']}."
            )
            rec = {
                "Input_ID": input_id,
                "Raw_Vehicle_Description": raw_desc,
                "Parsed_Brand": parsed_make,
                "Parsed_Model_Family": parsed_family,
                "Parsed_Year": parsed_year if parsed_year is not None else "",
                "Parsed_Generation": parsed_gen,
                "Input_Category": input_category,
                "Source": row["Support_List_Type"],
                "Matched_Brand": row["Brand"],
                "Matched_Model": row["Model"],
                "Matched_Vehicle_Category": row.get("Vehicle_Category",""),
                "Supported_Year_Range": row["Model_Year_Text"],
                "Brand_Check": score_info["brand_check"],
                "Model_Check": score_info["model_check"],
                "Year_Check": score_info["year_check"],
                "Category_Check": score_info["category_check"],
                "Supported_Parameter_Count": row.get("Supported_Parameter_Count",""),
                "Fuel_Data_Available": row.get("Fuel_Data_Available","No"),
                "Parameter_Preview": row.get("Param_Preview",""),
                "Full_Supported_Parameters": row.get("Full_Supported_Parameters",""),
                "Match_Score": score_info["score"],
                "Status": status,
                "Confidence": conf,
                "Reason": overall_reason(score_info["brand_check"], score_info["model_check"], score_info["year_check"], score_info["category_check"]),
                "Reason_Detail": reason_detail,
                "Year_Rank": score_info.get("year_rank", 0),
            }
            scored.append(rec)

        scored = sorted(scored, key=lambda x: (x["Match_Score"], x.get("Year_Rank", 0), {"High":3,"Medium":2,"Low":1}.get(x["Confidence"],0)), reverse=True)

        # anti-hallucination controls on top ranked candidates
        best_overall = None
        if scored:
            precise_scored = [c for c in scored if c.get("Year_Check") in {"Exact range", "Narrow open range"}]
            if precise_scored:
                scored = precise_scored + [c for c in scored if c not in precise_scored]

            top = scored[0]
            runner = scored[1] if len(scored) > 1 else None
            generic = generic_input_flag(parsed_family, search_text)
            same_family_conflict = runner and norm_text(top["Matched_Model"]) != norm_text(runner["Matched_Model"]) and abs(top["Match_Score"] - runner["Match_Score"]) < 6
            if not parsed_make:
                top["Status"] = "No Reliable Match"; top["Confidence"] = "Low"
            elif generic and top["Status"] == "Possible Match" and same_family_conflict:
                top["Status"] = "Review Needed"; top["Confidence"] = "Low"
            elif top["Year_Check"] == "Broad open range" and generic:
                top["Status"] = "Review Needed"; top["Confidence"] = "Low"
            elif top["Match_Score"] < 62 or top["Model_Check"] == "Weak":
                top["Status"] = "No Reliable Match"; top["Confidence"] = "Low"
            best_overall = top

        if not best_overall:
            best_overall = {
                "Parsed_Brand": parsed_make, "Parsed_Model_Family": parsed_family,
                "Parsed_Year": parsed_year or "", "Parsed_Generation": parsed_gen,
                "Input_Category": input_category,
                "Status":"No Reliable Match","Confidence":"Low","Source":"","Matched_Brand":"",
                "Matched_Model":"","Matched_Vehicle_Category":"","Supported_Year_Range":"",
                "Brand_Check":"No" if not parsed_make else "Exact","Model_Check":"Weak","Year_Check":"Not provided" if parsed_year is None else "",
                "Reason":"No candidates found","Reason_Detail":"No candidates found for the given input.",
                "Supported_Parameter_Count":"","Fuel_Data_Available":"No","Parameter_Preview":"","Full_Supported_Parameters":"","Match_Score":0
            }

        ws_match.cell(match_row_ptr, 1).value = input_id
        ws_match.cell(match_row_ptr, 2).value = raw_desc
        ws_match.cell(match_row_ptr, 3).value = best_overall["Parsed_Brand"]
        ws_match.cell(match_row_ptr, 4).value = best_overall["Parsed_Model_Family"]
        ws_match.cell(match_row_ptr, 5).value = best_overall["Parsed_Year"]
        ws_match.cell(match_row_ptr, 6).value = best_overall["Parsed_Generation"]
        ws_match.cell(match_row_ptr, 7).value = best_overall["Status"]
        ws_match.cell(match_row_ptr, 8).value = best_overall["Confidence"]
        ws_match.cell(match_row_ptr, 9).value = best_overall["Source"]
        ws_match.cell(match_row_ptr,10).value = best_overall["Matched_Brand"]
        ws_match.cell(match_row_ptr,11).value = best_overall["Matched_Model"]
        ws_match.cell(match_row_ptr,12).value = best_overall["Supported_Year_Range"]
        ws_match.cell(match_row_ptr,13).value = best_overall["Brand_Check"]
        ws_match.cell(match_row_ptr,14).value = best_overall["Model_Check"]
        ws_match.cell(match_row_ptr,15).value = best_overall["Year_Check"]
        ws_match.cell(match_row_ptr,16).value = best_overall["Reason"]
        ws_match.cell(match_row_ptr,17).value = action_hint(best_overall["Status"], parsed_year)
        ws_match.cell(match_row_ptr,18).value = best_overall["Supported_Parameter_Count"]
        ws_match.cell(match_row_ptr,19).value = best_overall["Fuel_Data_Available"]
        ws_match.cell(match_row_ptr,20).value = best_overall["Parameter_Preview"]
        match_row_ptr += 1

        for source in sorted({x["Source"] for x in scored if x.get("Source")}):
            candidates = [x for x in scored if x["Source"] == source]
            if not candidates:
                continue
            best = candidates[0]
            vals = [
                input_id, raw_desc, source, best["Status"], best["Confidence"],
                best["Parsed_Brand"], best["Parsed_Model_Family"], best["Parsed_Year"],
                best["Matched_Brand"], best["Matched_Model"], best["Supported_Year_Range"],
                best["Brand_Check"], best["Model_Check"], best["Year_Check"],
                best["Supported_Parameter_Count"], best["Fuel_Data_Available"], best["Parameter_Preview"], best["Reason"],
                best["Reason_Detail"], best["Full_Supported_Parameters"]
            ]
            for cidx, v in enumerate(vals, start=1):
                ws_source.cell(source_row_ptr, cidx).value = v
            source_row_ptr += 1

        for rank, cand in enumerate(scored[:3], start=1):
            vals = [
                input_id, raw_desc, rank, cand["Source"], cand["Status"], cand["Confidence"], cand["Match_Score"],
                cand["Matched_Brand"], cand["Matched_Model"], cand["Supported_Year_Range"],
                cand["Brand_Check"], cand["Model_Check"], cand["Year_Check"], cand["Reason"]
            ]
            for cidx, v in enumerate(vals, start=1):
                ws_top3.cell(top_row_ptr, cidx).value = v
            top_row_ptr += 1

        if scored:
            all_records.extend(scored)
        else:
            all_records.append({
                "Input_ID": input_id, "Raw_Vehicle_Description": raw_desc, "Parsed_Brand": parsed_make,
                "Parsed_Model_Family": parsed_family, "Parsed_Year": parsed_year if parsed_year is not None else "",
                "Parsed_Generation": parsed_gen, "Input_Category": input_category,
                "Source": "", "Matched_Brand": "", "Matched_Model": "", "Matched_Vehicle_Category": "",
                "Supported_Year_Range": "", "Brand_Check": "No" if not parsed_make else "Exact",
                "Model_Check": "Weak", "Year_Check": "Not provided" if parsed_year is None else "",
                "Category_Check": "Not provided", "Supported_Parameter_Count": "", "Fuel_Data_Available": "No",
                "Parameter_Preview": "", "Full_Supported_Parameters": "", "Match_Score": 0,
                "Status": best_overall["Status"], "Confidence": best_overall["Confidence"],
                "Reason": best_overall["Reason"], "Reason_Detail": best_overall["Reason_Detail"]
            })

        supported = "Yes" if best_overall["Status"] in ("Strong Match", "Possible Match") else "No"
        ws_in[f"D{r}"] = "Yes"
        ws_in[f"E{r}"] = f"Processed - {supported}"

    metrics, brand_counts, supported_df, unsupported_df, recommendations = build_focused_report(all_records, input_rows)

    color_results(ws_match, "G")
    color_results(ws_source, "D")
    color_results(ws_top3, "E")

    update_table_ref(ws_match, "T_MatchResults")
    update_table_ref(ws_source, "T_PerSourceResults")
    update_table_ref(ws_top3, "T_Top3Candidates")

    write_focused_report_sheets_to_workbook(wb, metrics, brand_counts, supported_df, unsupported_df, recommendations)
    wb.save(WORKBOOK)
    write_focused_report_workbook(metrics, brand_counts, supported_df, unsupported_df, recommendations)
    if not supported_df.empty:
        supported_df.to_csv(FINAL_REPORT_CSV, index=False)

    print(f"Processed {len(input_rows)} input row(s).")
    print(f"Per-source rows written: {max(0, source_row_ptr-2)}")
    print(f"Top-3 candidate rows written: {max(0, top_row_ptr-2)}")
    print(f"Smart report saved to: {FINAL_REPORT_XLSX.resolve()}")
    print(f"Workbook saved to: {WORKBOOK.resolve()}")

if __name__ == "__main__":
    main()
